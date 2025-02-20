# TO RUN: python3 app.py

# to download sheet as xlsx: use link: https://docs.google.com/spreadsheets/d/<KEY>/export?format=xlsx&gid=<GID>
# https://docs.google.com/spreadsheets/d/19UVHkup9MkuEqr_rVzx4sG6L1Hx3vr_g3O7HJrWXeu4/export?format=xlsx&gid=682182531
from flask import Flask, request, render_template, send_from_directory
import json
import os
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.styles import Color
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
import re
import math

import subprocess
import uuid
from pdf2image import convert_from_path

app = Flask(__name__)

# Set upload folder and allowed extensions
UPLOAD_FOLDER = 'uploads'
FILENAME = 'small_test.json'
TILE_FOLDER = 'tiles'
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['TILE_FOLDER'] = TILE_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)
if not os.path.exists(TILE_FOLDER):
    os.makedirs(TILE_FOLDER)

# Helper function to check allowed file types
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Upload route
@app.route('/')
def index():
    return '''
    <h1>Upload Excel File</h1>
    <form method="POST" action="/upload" enctype="multipart/form-data">
        <input type="file" name="file" accept=".xlsx, .xls">
        <input type="submit" value="Upload">
    </form>
    '''

@app.route('/upload', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return "No file part"
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        if file and allowed_file(file.filename):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file_id = "0"#str(uuid.uuid4())
            file.save(filepath)

            # Read the spreadsheet with pandas
            excel_data = pd.read_excel(filepath)

            # Convert HTML to PNG
            png_filename = os.path.splitext(file.filename)[0] + '.png'
            png_path = os.path.join(app.config['UPLOAD_FOLDER'], png_filename)

            # Process the sheet and get its information
            get_sheet_info(filepath)

            #libreoffice method
            #process_xlsx(filepath, file_id)

            if png_path:
                return f'png saved: <a href="{png_path}">{png_path}</a>'
            else:
                return "Failed to convert the file", 500

    return render_template('upload.html')

# Route to serve tiles
@app.route('/tiles/<int:zoom>/<int:x>/<int:y>.png')
def serve_tile(zoom, x, y):
    return send_from_directory(app.config['TILE_FOLDER'], f'{zoom}/{x}/{y}.png')

#--------------------------------------------------------------------------------------------------
#METHOD WHERE YOU DRAW LINES MANUALLY
#--------------------------------------------------------------------------------------------------
def get_sheet_info(filepath):
    wb = load_workbook(filepath)
    data_wb = load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames

    all_sheet_data = []
    sheet_index = 0
    grid_size = math.ceil(math.sqrt(len(sheet_names)))

    for sheet in sheet_names:
        ws = wb[sheet]

        # Find the last row and column with data
        max_row = ws.max_row
        max_column = ws.max_column

        # Scan each row and column to find the actual last used cell
        last_row = max((row for row in range(1, max_row + 1) if any(ws.cell(row=row, column=col).value is not None for col in range(1, max_column + 1))), default=1)
        last_column = max((col for col in range(1, max_column + 1) if any(ws.cell(row=row, column=col).value is not None for row in range(1, max_row + 1))), default=1)

        # Remove rows and columns beyond the last used cell
        if last_row < max_row:
            ws.delete_rows(last_row + 1, max_row - last_row)
        if last_column < max_column:
            ws.delete_cols(last_column + 1, max_column - last_column)

    max_width = max([get_sheet_width(wb[sheet]) for sheet in sheet_names])
    max_height = min([get_sheet_height(wb[sheet]) for sheet in sheet_names])
    print("HHHHHHHHHHHHHHHHHHh")
    for sheet in sheet_names:
        grid_row = sheet_index // grid_size
        grid_col = sheet_index % grid_size
        offset_x, offset_y = 0, 0
        for i in range(0, grid_col):
            offset_x += max_width + ((i + 1) * 50)
        for i in range(0, grid_row):
            offset_y += max_height - ((i + 1) * 50)
        ws = wb[sheet]
        data_ws = data_wb[sheet]

        print("sheet " + str(sheet_index))
        total_width = get_sheet_width(ws)
        total_height = get_sheet_height(ws)
        print("width: " + str(total_width))
        print("height: " + str(total_height))
        #print("height: " + str(total_height))
        sheet_data = {
            f"sheet{sheet_index}": {
                "rows": get_row_lines(ws, total_width, total_height, offset_x, offset_y),
                "cols": get_col_lines(ws, total_width, total_height, offset_x, offset_y),
                "cells": get_cells_info(ws, data_ws, sheet_index, offset_x, offset_y),
                "total_width": total_width,
                "total_height": total_height
            }
        }
        sheet_index += 1
        all_sheet_data.append(sheet_data)
    # Save to JSON
    sheet_info_path = os.path.join(UPLOAD_FOLDER, FILENAME)
    with open(sheet_info_path, "w") as f:
        json.dump(all_sheet_data, f, indent=4, default=str)

def get_merge_cells(ws):
    merged_cells = []
    start_cells = []
    for merged_range in ws.merged_cells.ranges:
        start = str(merged_range).split(":")[0]
        start_cells.append(start)
        for row in ws[merged_range.coord]:
            for cell in row:
                merged_cells.append(cell.coordinate)
    return start_cells, merged_cells

def get_row_lines(ws, w, h, x, y):
    rows = []
    y_offset = 0
    start_point = (0, 0)
    end_point = (0, 0)
    # Calculate row line positions
    for row in ws.iter_rows():
        row_index = row[0].row
        row_height = ws.row_dimensions[row_index].height or 15 # Default Excel row height
        start_point = (0 + x, y_offset + y)
        end_point = (w + x, y_offset + y)
        rows.append({"path": [start_point, end_point]})
        y_offset -= row_height
    #Do the last row too
    start_point = (0 + x, y_offset + y)
    end_point = (w + x, y_offset + y)
    rows.append({"path": [start_point, end_point]})
    return rows

def get_col_lines(ws, w, h, x, y):
    cols = []
    x_offset = 0
    start_point = (0, 0)
    end_point = (0, 0)
    # Calculate column line positions
    for col_letter in ws.column_dimensions:
        col_width = ws.column_dimensions[col_letter].width or 8.43  # Default width in characters
        col_width = col_width * 7.5  # Approx conversion from characters to points
        start_point = (x_offset + x, 0 + y)
        end_point = (x_offset + x, h + y)
        cols.append({"path": [start_point, end_point]})
        x_offset += col_width
    #Do the last col too
    start_point = (x_offset + x, 0 + y)
    end_point = (x_offset + x, h + y)
    cols.append({"path": [start_point, end_point]})
    return cols

formula_map = {}
def get_cells_info(ws, data_ws, index, x, y):
    cells = []
    x_offset = 0
    y_offset = 0

    # Track dependencies for all cells
    dependencies = {cell.coordinate: {'uses': [], 'used_by': [], 'weight': 0, 'rank': 0} for row in ws.iter_rows() for cell in row}

    # Recursively calculate weights for each cell
    def calculate_weight(cell_name):
        if dependencies[cell_name]["weight"] > 0:
            return dependencies[cell_name]["weight"], dependencies[cell_name]["rank"]
        
        weight = 0
        rank = -1
        
        for used_cell in dependencies[cell_name]['uses']:
            sub_weight, sub_rank = calculate_weight(used_cell)
            weight += 1 + sub_weight
            rank = max(rank, sub_rank)

        dependencies[cell_name]['weight'] = weight
        dependencies[cell_name]['rank'] = rank + 1
        
        return weight, rank + 1

    # Populate `uses` and `used_by` fields
    for row in ws.iter_rows():
        for cell in row:
            cell_name = cell.coordinate
            if cell.data_type == 'f':  # Cell contains a formula
                referenced_cells = extract_cell_references(cell.value)
                dependencies[cell_name]['uses'] = referenced_cells
                for ref in referenced_cells:
                    dependencies[ref]['used_by'].append(cell_name)

    for cell_name in dependencies:
        weight, rank = calculate_weight(cell_name)
        dependencies[cell_name]["weight"] = weight
        dependencies[cell_name]["rank"] = rank

    # Iterate through each cell and get its properties
    start_cells, merge_cells = get_merge_cells(ws)
    for row_num in range(1, ws.max_row + 1):
        row_height = ws.row_dimensions[row_num].height or 15  # Default height
        x_offset = 0
        for col in ws.iter_cols(min_row=row_num, max_row=row_num, min_col=1, max_col=ws.max_column):
            cell = col[0]
            cell_name = cell.coordinate
            col_width = 0
            
            
            if cell_name not in merge_cells or cell_name in start_cells:
                col_width = ws.column_dimensions[cell.column_letter].width or 8.43
                col_width = col_width * 7.5  # Convert width to points
            
            cell_value = ""
            if cell.data_type != 'f' and cell.value is not None:
                cell_value = cell.value
            elif cell.data_type == 'f':
                formula_type = extract_formula(cell.value, formula_map)
                cell_value = data_ws[cell_name].value
            cell_info = {
                "name": f"Sheet{index}!{cell_name}",
                "coord": (x_offset + x, y_offset + y),
                "height": row_height,
                "width": col_width,
                "value": clip_string_to_width(str(cell_value), col_width),
                "bg_color": get_color(cell, "fill"),
                "text_color": get_color(cell, "font"),
                "formula": cell.value if cell.data_type == 'f' else None,
                "formula_type": formula_map[formula_type] if cell.data_type == 'f' else None,
                "uses": [f"Sheet{index}!{cell_name}" for cell_name in dependencies[cell_name]['uses']],
                "used_by": [f"Sheet{index}!{cell_name}" for cell_name in dependencies[cell_name]['used_by']],
                "weight": dependencies[cell_name]['weight'],
                "rank": dependencies[cell_name]['rank']
            }
            cells.append(cell_info)
            x_offset += col_width
        y_offset -= row_height

    return cells

def get_sheet_width(ws):
    """Calculate the total width of the sheet based on actual cell data."""
    total_width = 0
    default_width = 8.43
    # Iterate through all columns (based on cell data)
    for col_num in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_num)
        col_width = ws.column_dimensions[col_letter].width or default_width
        col_width = col_width * 7.5  # Convert width to points
        total_width += col_width

    return total_width

def get_sheet_height(ws):
    """Calculate total height of the sheet based on row dimensions."""
    total_height = 0
    default_height = 15
    for row_num in range(1, ws.max_row + 1):
        row_height = ws.row_dimensions[row_num].height or default_height
        total_height -= row_height
    return total_height

def extract_formula(formula, formula_map):
    tokens = Tokenizer(formula)
    tokenless_formula = '='  # Start with '=' for all formulas

    # Parse each token in the formula
    for token in tokens.items:
        if token.type == 'OPERAND':  # Ignore cell and range references
            if token.subtype == 'RANGE' or token.subtype == 'CELL':
                tokenless_formula += ':'  # Placeholder for any range or cell reference
            elif token.subtype == 'NUMBER':  # Ignore numeric values
                tokenless_formula += 'n'
            else:
                tokenless_formula += token.value
        elif token.type == 'FUNCTION':  # Keep function names
            tokenless_formula += token.value
        elif token.type == 'OPERATOR-INFIX':  # Keep operators
            tokenless_formula += token.value
        elif token.type == 'PARENTHESIS':  # Keep parentheses
            tokenless_formula += token.value

    # Check if tokenless formula exists in dictionary
    if tokenless_formula not in formula_map:
        formula_map[tokenless_formula] = len(formula_map)  # Assign next integer value

    return tokenless_formula

def extract_cell_references(formula):
    """Extract individual cell references and expand ranges from a formula."""
    referenced_cells = []
    tokens = Tokenizer(formula).items  # Tokenize the formula         
    cell_references = []
    for t in tokens:
        if t.subtype == 'RANGE' or t.subtype == 'OPERAND':
            if ':' in t.value:  # It's a range, e.g., A1:B2
                min_col, min_row, max_col, max_row = range_boundaries(t.value)
                # Expand the range to individual cells
                for row in range(min_row, max_row + 1):
                    for col in range(min_col, max_col + 1):
                        cell_references.append(f"{get_column_letter(col)}{row}")
            elif any(char.isdigit() for char in t.value):
                # It's a single cell reference
                cell_references.append(re.sub(r'[^a-zA-Z0-9]', '', t.value))
    
    return cell_references

def get_color(cell, color_type="fill"):
    return [255, 255, 255, 255]
    color = None
    if color_type == 'fill':
        color = cell.fill.start_color.rgb
    elif color_type == 'text':
        color = cell.font.color.rgb
    # If color is None (no color set), return fully transparent (0, 0, 0, 0)
    if color is None or color == '00000000':  # No color or transparent
        return [255, 255, 255, 255]
    
    # Extract the RGB values (hex format is #RRGGBB)
    r = int(color[2:4], 16)
    g = int(color[4:6], 16)
    b = int(color[6:8], 16)
    a = int(color[0:2], 16)  # Alpha value (if available)
    
    return [r, g, b, a]

def clip_string_to_width(cell_value: str, cell_width: int) -> str:
    # Default character width in openpyxl (Calibri 11pt font size)
    char_width = 6  # approximate width of each character
    
    # Calculate the number of characters that can fit within the cell width
    max_chars = int(cell_width / char_width)
    
    # Split the string by newlines and process each line
    clipped_lines = []
    for line in cell_value.splitlines():
        if len(line) > max_chars:
            clipped_lines.append(line[:max_chars] + "...")
        else:
            clipped_lines.append(line)
    
    # Join the lines back with newlines
    return "\n".join(clipped_lines)


#--------------------------------------------------------------------------------------------------
#METHOD WHERE YOU GET A PDF/PNG OF SPREADSHEET
#--------------------------------------------------------------------------------------------------
def process_xlsx(file_path, map_id):
    # Create output directories
    map_tiles_dir = os.path.join(app.config['TILE_FOLDER'], map_id)
    os.makedirs(map_tiles_dir, exist_ok=True)

    temp_dir = os.path.join(app.config['UPLOAD_FOLDER'], map_id)
    os.makedirs(temp_dir, exist_ok=True)

    # Full path to the LibreOffice binary
    libreoffice_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    # Convert XLSX to PDF using LibreOffice
    cmd = [
        libreoffice_path,
        '--headless',  # No GUI
        '--convert-to', 'pdf:calc_pdf_Export:{"SinglePageSheets":{"type":"boolean","value":"true"}}',  # Conversion format
        '--outdir', temp_dir,  # Output directory
        '--calc',
        file_path  # Input file path
    ]
    subprocess.run(cmd, check=True)

    # Find the converted PDF file
    pdf_files = [f for f in os.listdir(temp_dir) if f.endswith('.pdf')]
    if not pdf_files:
        raise FileNotFoundError('PDF conversion failed.')
    pdf_path = os.path.join(temp_dir, pdf_files[0])

    # Convert PDF to images
    pages = convert_from_path(pdf_path)

    # Stitch all pages into a single image
    stitched_images = []
    total_height = 0
    max_width = 0

    for page in pages:
        stitched_images.append(page)
        total_height += page.height
        max_width = max(max_width, page.width)

    # Create a new blank image with the correct dimensions
    final_image = Image.new('RGB', (max_width, total_height), (255, 255, 255))

    # Paste each image into the final image
    current_y = 0
    for image in stitched_images:
        final_image.paste(image, (0, current_y))
        current_y += image.height

    # Save the final image temporarily
    final_image_path = os.path.join(temp_dir, 'final_image.png')
    final_image.save(final_image_path)

# Main entry point for the app
if __name__ == '__main__':
    app.run(debug=True)
